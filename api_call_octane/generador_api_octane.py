from PySide6.QtWidgets import QApplication, QLabel, QLineEdit, QMainWindow, QPushButton, QVBoxLayout, QWidget, QDialog
import openpyxl


class Formulario(QMainWindow):
    def __init__(self):
        super().__init__()

        # Crea los widgets de entrada de texto y las etiquetas
        self.octane_url_label = QLabel("Octane Url:")
        self.octane_url_text_field = QLineEdit()
        self.shared_space_label = QLabel("Shared Space:")
        self.shared_space_text_field = QLineEdit()
        self.udf_field_label = QLabel("UDF Field:")
        self.udf_field_text_field = QLineEdit()
        self.value_id_label_1 = QLabel("Deprecated value Id:")
        self.value_id_text_field_1 = QLineEdit()
        self.value_id_label_2 = QLabel("Deprecated value Id:")
        self.value_id_text_field_2 = QLineEdit()
        self.value_id_label_3 = QLabel("Deprecated value Id:")
        self.value_id_text_field_3 = QLineEdit()
        self.value_id_label_4 = QLabel("Deprecated value Id:")
        self.value_id_text_field_4 = QLineEdit()
        self.value_id_label_5 = QLabel("Deprecated value Id:")
        self.value_id_text_field_5 = QLineEdit()

        # Crea el botón "Exportar"
        self.exportar_button = QPushButton("Exportar")
        self.exportar_button.clicked.connect(self.exportar_datos)

        # Agrega los widgets de entrada de texto y las etiquetas a un layout vertical
        layout = QVBoxLayout()
        layout.addWidget(self.octane_url_label)
        layout.addWidget(self.octane_url_text_field)
        layout.addWidget(self.shared_space_label)
        layout.addWidget(self.shared_space_text_field)
        layout.addWidget(self.udf_field_label)
        layout.addWidget(self.udf_field_text_field)
        layout.addWidget(self.value_id_label_1)
        layout.addWidget(self.value_id_text_field_1)
        layout.addWidget(self.value_id_label_2)
        layout.addWidget(self.value_id_text_field_2)
        layout.addWidget(self.value_id_label_3)
        layout.addWidget(self.value_id_text_field_3)
        layout.addWidget(self.value_id_label_4)
        layout.addWidget(self.value_id_text_field_4)
        layout.addWidget(self.value_id_label_5)
        layout.addWidget(self.value_id_text_field_5)


        # Agrega el botón "Exportar" al layout
        layout.addWidget(self.exportar_button)
        

        # Crea un widget central que contenga el layout vertical
        widget = QWidget()
        widget.setLayout(layout)

        # Establece el widget central de la ventana como el widget que acabamos de crear
        self.setCentralWidget(widget)

    def mostrarError(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Please type at least a deprecated value!")
        dlg.exec_()
    
    #itera en todos los QLineEdit y devuelve un array con el texto de estos
    #ERROR HAY TRES TEXTBOXES QUE NO SE TIENEN QUE TOMAR EN CUENTA  
    def get_textFromTextBoxes(self):
        texts = []
        for line_edit in self.findChildren(QLineEdit):
            if line_edit.text():
                text = line_edit.text()
                texts.append(text)
        return texts


    def exportar_datos(self):
        # Obtener los valores de entrada de texto
        octane_url = self.octane_url_text_field.text()
        shared_space = self.shared_space_text_field.text()
        udf_field = self.udf_field_text_field.text()
        entity = "work_items"
        deprecated_ids = ""

        #Si hay multiples IDs los pone todos en una lista
        widgets = self.get_textFromTextBoxes()#deprecated values will be saved here
        for text in widgets:
            #si es textbox y no esta vacio
            if len(deprecated_ids) == 0:
                deprecated_ids = "'" + text + "'"
            else:
                deprecated_ids = ",'" + text + "'"

        print(deprecated_ids)
       # if len(deprecated_ids) != 0
        # Abre el archivo Excel donde se copian los workspaces
        workbook = openpyxl.load_workbook('WORKSPACES.xlsx')

        # Selecciona la hoja del archivo
        worksheet = workbook['WORKSPACES']

        

        if len(deprecated_ids) == 0: 
            self.mostrarError()

        else:
            # Lee los workspaces de la columna A y los almacena en una lista
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                

                wsid = row[0]
                #name = row[1]
                #status = row[2]

                #concatenated_string = f"{column_a_value} {column_b_value}"
                APIcall = octane_url + "/api/shared_spaces/" + str(shared_space) + "/workspaces/" + str(wsid) + "/" + entity + "?fields=id,name," + udf_field + "&query=\"" + udf_field + "={id+IN+'" + str(deprecated_ids) + "'}\""
                #print(APIcall) #test purposes
                data.append(APIcall)
            
            

            # Crear un nuevo libro de Excel
            new_workbook = openpyxl.Workbook()

            # Crear una nueva hoja de trabajo en el libro de Excel
            new_worksheet = new_workbook.active
            new_worksheet.title = 'workitems'

            # Escribir los datos de la estructura de datos junto con los valores de entrada de texto en el archivo de Excel
            for row in data:
                #for column in row:
                    #new_worksheet.cell(row=row.index(column) + 1, column=1, value=str(column))
                #print(row.index(row)) # testing
                new_worksheet.cell(data.index(row)+1, 1, row)

            # Guardar el libro de Excel
            new_workbook.save("resultados.xlsx")



if __name__ == '__main__':
   
    app = QApplication([])
    ventana = Formulario()
    ventana.show()
    app.exec_()
